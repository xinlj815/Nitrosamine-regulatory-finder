#!/usr/bin/env python3
"""Refresh nitrosamine AI limits from official regulator sources.

The script is designed for GitHub Actions. It:
1. Downloads and parses EMA, Health Canada, FDA and TGA data.
2. Normalises CAS/name records.
3. Keeps the previous regulator data if one source is temporarily unavailable.
4. Writes a change log.
5. Optionally sends changes to the user's existing Cloudflare -> DingTalk relay.

Environment variables:
  DINGTALK_RELAY_URL
  DINGTALK_RELAY_TOKEN
  EMA_XLSX_PATH / HC_XLSX_PATH / FDA_HTML_PATH
  TGA_CSV_PATH / TGA_HTML_PATH / TGA_CSV_URL  (offline/URL overrides)
"""
from __future__ import annotations

import argparse
import copy
import datetime as dt
import json
import os
import re
import sys
import tempfile
import time
import unicodedata
from dataclasses import dataclass
from io import StringIO
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA_FILE = ROOT / "data" / "nitrosamine_limits.json"
CHANGES_FILE = ROOT / "data" / "changes.json"
ALIASES_FILE = ROOT / "data" / "manual_aliases.json"

EMA_URL = "https://www.ema.europa.eu/en/documents/other/appendix-1-acceptable-intakes-established-n-nitrosamines_en.xlsx"
HC_URL = "https://www.canada.ca/content/dam/hc-sc/documents/services/drugs-health-products/compliance-enforcement/information-health-product/drugs/nitrosamine-impurities/established-acceptable-intake-limits/appendix-1.xlsx"
FDA_URL = "https://www.fda.gov/regulatory-information/search-fda-guidance-documents/cder-nitrosamine-impurity-acceptable-intake-limits"
TGA_URL = "https://www.tga.gov.au/safety/safety-monitoring-and-information/nitrosamine-impurities-medicines/established-acceptable-intake-nitrosamines-medicines"
TGA_CSV_FILENAME = "Established_acceptable_intake_for_nitrosamines.csv"

REQUEST_TIMEOUT = (20, 180)
TRANSIENT_HTTP_STATUS = {429, 500, 502, 503, 504}

# TGA intermittently stalls for GitHub-hosted runners.  Keep these limits
# deliberately short: two CSV attempts plus one HTML fallback finish in about
# one minute when the route is unavailable, after which the last good TGA data
# is retained by --allow-partial.
TGA_CSV_ATTEMPTS = 2
TGA_CSV_TIMEOUT = (8, 20)
TGA_HTML_ATTEMPTS = 1
TGA_HTML_TIMEOUT = (8, 25)

AGENCIES = ("EMA", "Health Canada", "FDA", "TGA")
SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "NitrosamineRegulatoryFinder/1.0 (+GitHub Actions; regulatory data monitor)",
    "Accept-Language": "en",
})


def get_with_retry(
    url: str,
    *,
    attempts: int = 4,
    timeout: tuple[int, int] = REQUEST_TIMEOUT,
) -> requests.Response:
    """GET an official source with bounded retries for transient failures."""
    last_error: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            response = SESSION.get(url, timeout=timeout)
            if response.status_code in TRANSIENT_HTTP_STATUS:
                raise requests.HTTPError(
                    f"transient HTTP {response.status_code}", response=response
                )
            response.raise_for_status()
            return response
        except (requests.Timeout, requests.ConnectionError, requests.HTTPError) as exc:
            status = getattr(getattr(exc, "response", None), "status_code", None)
            if isinstance(exc, requests.HTTPError) and status not in TRANSIENT_HTTP_STATUS:
                raise
            last_error = exc
            if attempt == attempts:
                break
            delay = min(2 ** (attempt - 1), 8)
            print(
                f"[RETRY] GET {url} failed ({attempt}/{attempts}): "
                f"{type(exc).__name__}: {exc}; retrying in {delay}s",
                file=sys.stderr,
            )
            time.sleep(delay)
    assert last_error is not None
    raise last_error


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKD", clean(value)).lower()
    text = text.translate(str.maketrans({"‐": "-", "‑": "-", "‒": "-", "–": "-", "—": "-", "―": "-", "′": "'"}))
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", text)


def name_variants(value: Any) -> set[str]:
    """Return conservative name variants for cross-regulator matching.

    FDA often appends an acronym in parentheses, for example
    ``N-nitroso-dimethylamine (NDMA)``, whereas EMA may publish the same
    compound without the acronym.  Only acronym/alias-like parenthetical
    groups are removed; stereochemical markers such as (R) and (S) are kept
    to avoid merging stereoisomers.
    """
    text = clean(value)
    if not text:
        return set()

    variants = {normalize(text)}

    def strip_alias_group(match: re.Match[str]) -> str:
        group = clean(match.group(1))
        acronym = bool(re.fullmatch(r"[A-Z0-9][A-Z0-9\-_/ ]{1,29}", group))
        alias_phrase = group.lower().startswith(("aka ", "also known as ", "or "))
        return " " if acronym or alias_phrase else match.group(0)

    without_alias_parentheses = re.sub(r"\(([^()]*)\)", strip_alias_group, text)
    variants.add(normalize(without_alias_parentheses))

    # FDA occasionally uses "aka" in the official display name.  Treat both
    # sides as searchable variants without changing the displayed source name.
    for part in re.split(r"\s+(?:aka|also known as)\s+", without_alias_parentheses, flags=re.I):
        if clean(part):
            variants.add(normalize(part))

    return {item for item in variants if item}


def valid_cas(value: Any) -> bool:
    return bool(re.fullmatch(r"\d{2,7}-\d{2}-\d", clean(value)))


def parse_number(value: Any) -> float | None:
    text = clean(value).replace(",", "")
    match = re.match(r"^(\d+(?:\.\d+)?)", text)
    return float(match.group(1)) if match else None


def iso_date(value: Any) -> str:
    if value is None or clean(value) == "":
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.date().isoformat() if isinstance(value, dt.datetime) else value.isoformat()
    text = clean(value)
    # ISO already
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return text
    # Excel numeric date
    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        date = dt.datetime(1899, 12, 30) + dt.timedelta(days=float(text))
        return date.date().isoformat()
    # Try common natural date formats
    for fmt in ("%d %B %Y", "%B %d, %Y", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def source_status(version: str, updated: str, url: str, mode: str = "full") -> dict[str, str]:
    return {"version": version, "updated": updated, "url": url, "mode": mode}


def blank_record(name: str, cas: str = "") -> dict[str, Any]:
    return {
        "id": "",
        "name": clean(name),
        "cas": clean(cas) if valid_cas(cas) else "",
        "aliases": [],
        "related_substances": [],
        "smiles": "",
        "iupac": "",
        "regulators": {},
        "special_limits": [],
    }


class Registry:
    def __init__(self, aliases: dict[str, list[str]]):
        self.records: dict[str, dict[str, Any]] = {}
        self.aliases = aliases

    @staticmethod
    def key(name: str, cas: str = "") -> str:
        return f"cas:{clean(cas)}" if valid_cas(cas) else f"name:{normalize(name)}"

    def find_by_name(self, name: str) -> dict[str, Any] | None:
        needles = name_variants(name)
        if not needles:
            return None
        for record in self.records.values():
            candidates = name_variants(record["name"])
            for alias in record["aliases"]:
                candidates.update(name_variants(alias))
            if needles & candidates:
                return record
        return None

    def get(self, name: str, cas: str = "", aliases: Iterable[str] = ()) -> dict[str, Any]:
        name, cas = clean(name), clean(cas)
        if not valid_cas(cas):
            cas = ""
        key = self.key(name, cas)

        # Prefer an existing CAS record, then exact normalised-name record.
        record = self.records.get(key)
        if record is None and cas:
            record = next((x for x in self.records.values() if x.get("cas") == cas), None)
        if record is None:
            record = self.find_by_name(name)

        if record is None:
            record = blank_record(name, cas)
            self.records[key] = record
        elif key not in self.records:
            # Record was found through name; keep one object under its original key.
            pass

        if cas and not record["cas"]:
            record["cas"] = cas
        for alias in aliases:
            alias = clean(alias)
            if alias and normalize(alias) != normalize(record["name"]) and alias not in record["aliases"]:
                record["aliases"].append(alias)
        return record

    def add_regulator(
        self,
        *,
        agency: str,
        name: str,
        cas: str = "",
        aliases: Iterable[str] = (),
        related: Iterable[str] = (),
        ai_raw: Any = "",
        cpca: Any = "",
        basis: str = "",
        publication_date: Any = "",
        source_url: str,
        source_version: str,
        smiles: str = "",
        iupac: str = "",
        status: str | None = None,
        source_table: str = "",
    ) -> None:
        record = self.get(name, cas, aliases)
        if smiles and not record["smiles"]:
            record["smiles"] = clean(smiles)
        if iupac and not record["iupac"]:
            record["iupac"] = clean(iupac)
        for related_name in related:
            related_name = clean(related_name)
            if related_name and related_name not in ("-", "_", "―", "NA", "N/A") and related_name not in record["related_substances"]:
                record["related_substances"].append(related_name)

        ai_display = clean(ai_raw) or "未建立数值AI"
        ai = parse_number(ai_display)
        if status is None:
            upper = ai_display.upper()
            if ai is not None:
                status = "numeric"
            elif "NMI" in upper or "**" in upper:
                status = "nmi"
            else:
                status = "other"

        current = record["regulators"].get(agency)
        new_value = {
            "ai_ng_day": ai,
            "ai_display": ai_display,
            "cpca_category": clean(cpca).replace("―", ""),
            "basis": clean(basis),
            "publication_date": iso_date(publication_date),
            "source_url": source_url,
            "source_version": source_version,
            "status": status,
            "source_table": clean(source_table),
        }
        if current:
            # Same compound may appear several times for different APIs. Preserve related APIs
            # and prefer a numeric, latest/compound-specific value over an empty one.
            if current.get("ai_ng_day") is not None and new_value["ai_ng_day"] is None:
                return
            if current.get("ai_ng_day") == new_value["ai_ng_day"]:
                if new_value["publication_date"] > current.get("publication_date", ""):
                    current["publication_date"] = new_value["publication_date"]
                if new_value["basis"] and new_value["basis"] not in current.get("basis", ""):
                    current["basis"] = "; ".join(filter(None, [current.get("basis", ""), new_value["basis"]]))
                if new_value["source_table"] and new_value["source_table"] not in current.get("source_table", ""):
                    current["source_table"] = "; ".join(filter(None, [current.get("source_table", ""), new_value["source_table"]]))
                return
        record["regulators"][agency] = new_value

    def add_special_limit(
        self,
        *,
        agency: str,
        limit_type: str,
        name: str,
        cas: str = "",
        aliases: Iterable[str] = (),
        related: Iterable[str] = (),
        ai_raw: Any = "",
        official_control_raw: Any = "",
        applicable_product: str = "",
        estimated_duration: Any = "",
        basis: str = "",
        publication_date: Any = "",
        source_url: str,
        source_version: str,
        source_table: str = "",
    ) -> None:
        record = self.get(name, cas, aliases)
        for related_name in related:
            related_name = clean(related_name)
            if related_name and related_name not in ("-", "_", "―", "NA", "N/A") and related_name not in record["related_substances"]:
                record["related_substances"].append(related_name)

        ai_display = clean(ai_raw) or "未建立数值AI"
        control_display = clean(official_control_raw)
        if control_display and "ppm" not in control_display.lower() and parse_number(control_display) is not None:
            control_display = f"{control_display} ppm"
        item = {
            "agency": clean(agency),
            "limit_type": clean(limit_type),
            "ai_ng_day": parse_number(ai_display),
            "ai_display": ai_display,
            "official_control_ppm": parse_number(control_display),
            "official_control_display": control_display,
            "applicable_product": clean(applicable_product),
            "estimated_duration": iso_date(estimated_duration),
            "basis": clean(basis),
            "publication_date": iso_date(publication_date),
            "source_url": source_url,
            "source_version": source_version,
            "source_table": clean(source_table),
            "status": clean(limit_type) or "special",
        }
        identity = (
            normalize(item["agency"]),
            normalize(item["limit_type"]),
            normalize(item["applicable_product"]),
        )
        for index, current in enumerate(record["special_limits"]):
            current_identity = (
                normalize(current.get("agency")),
                normalize(current.get("limit_type")),
                normalize(current.get("applicable_product")),
            )
            if current_identity == identity:
                record["special_limits"][index] = item
                return
        record["special_limits"].append(item)

    def apply_aliases(self) -> None:
        for cas, aliases in self.aliases.items():
            record = next((x for x in self.records.values() if x.get("cas") == cas), None)
            if not record:
                continue
            for alias in aliases:
                if alias not in record["aliases"]:
                    record["aliases"].append(alias)

    def serialise(self) -> list[dict[str, Any]]:
        output = []
        for record in self.records.values():
            key = self.key(record["name"], record["cas"])
            record["id"] = key.replace(":", "-")
            record["aliases"] = sorted(set(filter(None, map(clean, record["aliases"]))), key=str.lower)
            record["related_substances"] = sorted(set(filter(None, map(clean, record["related_substances"]))), key=str.lower)
            record["special_limits"] = sorted(
                record.get("special_limits", []),
                key=lambda x: (x.get("agency", ""), x.get("limit_type", ""), x.get("applicable_product", "")),
            )
            output.append(record)
        return sorted(output, key=lambda x: (x["name"].lower(), x["cas"]))


def download(url: str, suffix: str) -> Path:
    response = get_with_retry(url)
    handle = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    handle.write(response.content)
    handle.close()
    return Path(handle.name)


def find_header(sheet, required_terms: Iterable[str]) -> tuple[int, dict[str, int]]:
    terms = [normalize(x) for x in required_terms]
    for row in sheet.iter_rows():
        values = [clean(cell.value) for cell in row]
        joined = " ".join(values)
        if all(term in normalize(joined) for term in terms):
            mapping = {normalize(value): idx for idx, value in enumerate(values) if value}
            return row[0].row, mapping
    raise ValueError(f"Header not found: {required_terms}")


def parse_ema(path: Path, registry: Registry) -> dict[str, str]:
    """Parse both worksheets in EMA Appendix 1.

    Rev. 13 contains a primary ``N-nitrosamines`` worksheet and a second
    ``Other N-nitroso-structures`` worksheet.  The previous implementation
    only opened the first worksheet, which omitted the second regulatory table.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = [
        name for name in ("N-nitrosamines", "Other N-nitroso-structures")
        if name in wb.sheetnames
    ]
    if not sheet_names:
        sheet_names = wb.sheetnames

    first_ws = wb[sheet_names[0]]
    metadata = [
        clean(cell)
        for row in first_ws.iter_rows(min_row=1, max_row=12, values_only=True)
        for cell in row
        if clean(cell)
    ]
    version_text = next((x for x in metadata if x.startswith("EMA/")), "EMA Appendix 1")
    date_match = next(
        (re.search(r"(\d{4}-\d{2}-\d{2})", x) for x in metadata if re.search(r"\d{4}-\d{2}-\d{2}", x)),
        None,
    )
    document_date = date_match.group(1) if date_match else ""

    counts: dict[str, int] = {}
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        header_row, _ = find_header(ws, ("Name", "CAS RN", "AI"))
        headers = [clean(cell.value) for cell in ws[header_row]]
        index = {normalize(value): i for i, value in enumerate(headers) if value}

        def col(*candidates: str, required: bool = True) -> int | None:
            for candidate in candidates:
                nc = normalize(candidate)
                for key, idx in index.items():
                    if nc == key or nc in key:
                        return idx
            if required:
                raise KeyError(candidates)
            return None

        i_name = col("Name")
        i_iupac = col("IUPAC name")
        i_smiles = col("SMILES")
        i_cas = col("CAS RN")
        i_alias = col("Synonym Acronym")
        i_source = col("Source")
        i_cpca = col("CPCA Category", required=False)
        i_ai = col("AI ng day")
        i_note = col("Note")
        i_date = col("Publication date")

        count = 0
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            name = clean(row[i_name])
            if not name:
                continue
            aliases = [x.strip() for x in re.split(r"[,/;]", clean(row[i_alias])) if x.strip()]
            registry.add_regulator(
                agency="EMA",
                name=name,
                cas=clean(row[i_cas]),
                aliases=aliases,
                related=[row[i_source]],
                ai_raw=row[i_ai],
                cpca=row[i_cpca] if i_cpca is not None else "",
                basis=clean(row[i_note]),
                publication_date=row[i_date],
                source_url=EMA_URL,
                source_version=version_text,
                source_table=f"EMA Appendix 1 - {sheet_name}",
                smiles=clean(row[i_smiles]),
                iupac=clean(row[i_iupac]),
            )
            count += 1
        counts[sheet_name] = count

    count_text = "; ".join(f"{name}: {count} rows" for name, count in counts.items())
    return source_status(
        f"{version_text}; {count_text}",
        document_date or dt.date.today().isoformat(),
        EMA_URL,
        mode="both Appendix 1 worksheets",
    )

def parse_health_canada(path: Path, registry: Registry) -> dict[str, str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    metadata = [clean(cell) for row in ws.iter_rows(min_row=1, max_row=15, values_only=True) for cell in row if clean(cell)]
    version_match = next((re.search(r"version:\s*(\d{4}-\d{2}-\d{2})", x, flags=re.I) for x in metadata if "version" in x.lower()), None)
    version_date = version_match.group(1) if version_match else ""
    version_text = f"Health Canada Appendix 1; version {version_date}" if version_date else "Health Canada Appendix 1"

    header_row, _ = find_header(ws, ("N-nitrosamine", "CAS RN", "AI Limit"))
    headers = [clean(cell.value) for cell in ws[header_row]]
    index = {normalize(value): i for i, value in enumerate(headers) if value}

    def col(part: str) -> int:
        npart = normalize(part)
        return next(idx for key, idx in index.items() if npart in key)

    i_date = col("Date Most Recently Published")
    i_related = col("Related Drug Substance")
    i_name = col("N-nitrosamine")
    i_cas = col("CAS RN")
    i_cpca = col("CPCA Category")
    i_ai = col("AI Limit")

    count = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        name = clean(row[i_name])
        if not name:
            continue
        raw_ai = clean(row[i_ai])
        notes = []
        if "**" in raw_ai:
            notes.append("按 ICH Q3A/Q3B 作为非致突变杂质控制")
        if "***" in raw_ai:
            notes.append("基于阴性体外细菌回复突变试验")
        if "^" in raw_ai:
            notes.append("证据权重法")
        aliases: list[str] = []
        for group in re.findall(r"\(([^()]*)\)", name):
            if len(group) <= 30 and not valid_cas(group):
                aliases.extend(x.strip() for x in re.split(r"[,/]", group) if 1 < len(x.strip()) <= 20)
        registry.add_regulator(
            agency="Health Canada", name=name, cas=clean(row[i_cas]), aliases=aliases,
            related=[row[i_related]], ai_raw=raw_ai, cpca=row[i_cpca],
            basis="；".join(notes), publication_date=row[i_date],
            source_url=HC_URL, source_version=version_text,
            source_table="Health Canada Appendix 1"
        )
        count += 1
    return source_status(f"{version_text}; {count} rows", version_date or dt.date.today().isoformat(), HC_URL)


def flatten_columns(frame: pd.DataFrame) -> pd.DataFrame:
    if isinstance(frame.columns, pd.MultiIndex):
        frame.columns = [" ".join(clean(x) for x in col if clean(x)) for col in frame.columns]
    else:
        frame.columns = [clean(x) for x in frame.columns]
    return frame


def find_column(frame: pd.DataFrame, *parts: str) -> str:
    normalized = {column: normalize(column) for column in frame.columns}
    for part in parts:
        npart = normalize(part)
        for column, ncolumn in normalized.items():
            if npart in ncolumn:
                return column
    raise KeyError(parts)


def parse_fda(registry: Registry) -> dict[str, str]:
    """Parse FDA Tables 1, 2 and 3.

    Tables 1 and 2 contain the regular recommended AI limits. Table 3 contains
    temporary, product-specific interim AI/control limits. Interim values are
    stored separately so they never overwrite the regular lifetime AI.
    """
    offline_path = os.getenv("FDA_HTML_PATH", "").strip()
    if offline_path:
        html_text = Path(offline_path).read_text(encoding="utf-8")
    else:
        response = get_with_retry(FDA_URL)
        html_text = response.text

    soup = BeautifulSoup(html_text, "lxml")

    def section_updated(table_label: str) -> str:
        heading = soup.find(
            lambda tag: getattr(tag, "name", None) in {"h2", "h3", "h4"}
            and table_label.lower() in clean(tag.get_text(" ", strip=True)).lower()
        )
        if heading:
            for node in heading.find_all_next():
                if node is not heading and getattr(node, "name", None) in {"h2", "h3", "h4"}:
                    break
                text = clean(node.get_text(" ", strip=True)) if getattr(node, "get_text", None) else clean(node)
                match = re.search(r"Updated:\s*(\d{1,2}/\d{1,2}/\d{4})", text, flags=re.I)
                if match:
                    return iso_date(match.group(1))
        return ""

    table3_updated = section_updated("Table 3")
    tables = [flatten_columns(frame) for frame in pd.read_html(StringIO(html_text))]
    regular_parsed = 0
    interim_parsed = 0
    interim_table_seen = False

    for frame in tables:
        normalized_columns = {column: normalize(column) for column in frame.columns}
        has_name = any("nitrosaminename" in value for value in normalized_columns.values())
        if not has_name:
            continue

        is_interim = any("recommendedinterimailimit" in value for value in normalized_columns.values())
        is_regular = any(
            "recommendedailimit" in value and "interim" not in value
            for value in normalized_columns.values()
        )

        if is_interim:
            interim_table_seen = True
            name_col = find_column(frame, "Nitrosamine Name")
            source_col = find_column(frame, "Source")
            ai_col = find_column(frame, "Recommended Interim AI Limit")
            ppm_col = find_column(frame, "Recommended Interim Control Limit")
            duration_col = find_column(frame, "Estimated Duration")

            for _, row in frame.iterrows():
                name = clean(row.get(name_col))
                if not name or name.lower() == "nan":
                    continue
                product = clean(row.get(source_col))
                registry.add_special_limit(
                    agency="FDA",
                    limit_type="interim",
                    name=name,
                    related=[product],
                    ai_raw=row.get(ai_col),
                    official_control_raw=row.get(ppm_col),
                    applicable_product=product,
                    estimated_duration=row.get(duration_col),
                    basis=(
                        "FDA Table 3临时AI：仅适用于表中所列已批准/在售产品，"
                        "用于潜在供应中断情形下的阶段性监管安排；不替代常规终生AI。"
                    ),
                    publication_date=table3_updated,
                    source_url=FDA_URL,
                    source_version="FDA Table 3 - Recommended Interim AI Limits",
                    source_table="FDA Table 3",
                )
                interim_parsed += 1
            continue

        if not is_regular:
            continue

        name_col = find_column(frame, "Nitrosamine Name")
        ai_col = find_column(frame, "Recommended AI Limit")
        source_col = next((c for c in frame.columns if "source" in normalize(c)), "")
        potency_col = next((c for c in frame.columns if "potencycategory" in normalize(c)), "")
        surrogate_col = next((c for c in frame.columns if "surrogate" in normalize(c)), "")
        date_col = next((c for c in frame.columns if "dateadded" in normalize(c)), "")
        table_name = "FDA Table 1" if potency_col else "FDA Table 2"

        for _, row in frame.iterrows():
            name = clean(row.get(name_col))
            if not name or name.lower() == "nan":
                continue
            ai_raw = clean(row.get(ai_col))
            basis = clean(row.get(surrogate_col)) if surrogate_col else ""
            if not basis:
                basis = "FDA CPCA" if potency_col else "FDA compound-specific/read-across table"
            registry.add_regulator(
                agency="FDA",
                name=name,
                related=[row.get(source_col)] if source_col else [],
                ai_raw=ai_raw,
                cpca=row.get(potency_col) if potency_col else "",
                basis=basis,
                publication_date=row.get(date_col) if date_col else "",
                source_url=FDA_URL,
                source_version=f"{table_name} (current page)",
                source_table=table_name,
            )
            regular_parsed += 1

    text = soup.get_text(" ", strip=True)
    dates = re.findall(r"Updated:\s*(\d{1,2}/\d{1,2}/\d{4})", text)
    latest = max((iso_date(x) for x in dates), default=dt.date.today().isoformat())
    if regular_parsed == 0:
        raise ValueError("No FDA Table 1/2 AI rows parsed")
    if not interim_table_seen:
        raise ValueError("FDA Table 3 interim AI table not found")
    return source_status(
        f"FDA online AI tables; {regular_parsed} regular rows; {interim_parsed} interim rows",
        latest,
        FDA_URL,
        mode="Tables 1, 2 and 3",
    )

def split_tga_name(cell: str) -> tuple[str, str, list[str]]:
    text = clean(cell)
    cas_match = re.search(r"(\d{2,7}-\d{2}-\d)", text)
    cas = cas_match.group(1) if cas_match else ""
    aliases: list[str] = []
    for group in re.findall(r"\(([^()]*)\)", text):
        if cas and cas in group:
            remainder = group.replace(cas, "")
            aliases.extend(x.strip(" ,/") for x in re.split(r"[,/]", remainder) if x.strip(" ,/"))
        elif len(group) <= 30:
            aliases.extend(x.strip() for x in re.split(r"[,/]", group) if x.strip())
    # Use the first line/first slash component as the concise display name and
    # retain any short alternative name as an alias.
    name_parts = [clean(x) for x in re.split(r"\s*/\s*|\n", text) if clean(x)]
    name = name_parts[0] if name_parts else text
    aliases.extend(x for x in name_parts[1:] if len(x) <= 120)
    if cas:
        name = name.replace(cas, "").strip(" ()-,")
    return name, cas, aliases


def split_tga_identifiers(value: Any) -> tuple[str, list[str]]:
    """Split TGA's separate 'CAS No, abbreviations' field."""
    text = clean(value)
    cas_values = re.findall(r"\b\d{2,7}-\d{2}-\d\b", text)
    cas = cas_values[0] if cas_values else ""
    remainder = text
    for item in cas_values:
        remainder = remainder.replace(item, " ")
    aliases = [
        token.strip(" ()")
        for token in re.split(r"[,;/\n]", remainder)
        if token.strip(" ()")
    ]
    return cas, aliases


def latest_tga_date(*values: Any) -> str:
    dates: list[str] = []
    for value in values:
        for part in re.split(r"[;,\n]", clean(value)):
            parsed = iso_date(part.strip())
            if re.fullmatch(r"\d{4}-\d{2}-\d{2}", parsed):
                dates.append(parsed)
    return max(dates, default="")


def tga_csv_candidates(months: int = 18) -> list[str]:
    """Return newest-first TGA CSV URLs, allowing future monthly updates."""
    override = os.getenv("TGA_CSV_URL", "").strip()
    if override:
        return [override]
    today = dt.date.today()
    year, month = today.year, today.month
    urls: list[str] = []
    for _ in range(months):
        urls.append(
            f"https://www.tga.gov.au/sites/default/files/{year:04d}-{month:02d}/"
            f"{TGA_CSV_FILENAME}"
        )
        month -= 1
        if month == 0:
            year -= 1
            month = 12
    return urls


def parse_tga(registry: Registry) -> dict[str, str]:
    """Parse TGA AI data, preferring its small official CSV download.

    TGA's HTML page is large and intermittently times out from GitHub-hosted
    runners.  The official CSV contains the same table and is much more stable;
    the HTML parser remains as a fallback for future URL changes.
    """
    offline_csv = os.getenv("TGA_CSV_PATH", "").strip()
    offline_html = os.getenv("TGA_HTML_PATH", "").strip()
    frame: pd.DataFrame | None = None
    mode = "official CSV"
    page_updated = ""
    csv_error: Exception | None = None

    try:
        if offline_csv:
            csv_text = Path(offline_csv).read_text(encoding="utf-8-sig")
        else:
            csv_response: requests.Response | None = None
            candidate_errors: list[str] = []
            for candidate in tga_csv_candidates():
                try:
                    csv_response = get_with_retry(
                        candidate,
                        attempts=TGA_CSV_ATTEMPTS,
                        timeout=TGA_CSV_TIMEOUT,
                    )
                    break
                except requests.HTTPError as exc:
                    if getattr(exc.response, "status_code", None) == 404:
                        candidate_errors.append(f"{candidate}: 404")
                        continue
                    raise
            if csv_response is None:
                raise FileNotFoundError(
                    "No recent TGA CSV URL was found; " + "; ".join(candidate_errors)
                )
            try:
                csv_text = csv_response.content.decode("utf-8-sig")
            except UnicodeDecodeError:
                csv_text = csv_response.text.lstrip("\ufeff")
        frame = flatten_columns(
            pd.read_csv(StringIO(csv_text), dtype=str, keep_default_na=False)
        )
        if not any("nitrosamine" in normalize(c) for c in frame.columns) or not any(
            "ailimit" in normalize(c) for c in frame.columns
        ):
            raise ValueError("TGA CSV headers not recognised")
    except Exception as exc:
        csv_error = exc
        print(
            f"[WARN] TGA CSV unavailable ({type(exc).__name__}: {exc}); "
            "falling back to the official HTML page",
            file=sys.stderr,
        )
        mode = "HTML fallback"
        if offline_html:
            html_text = Path(offline_html).read_text(encoding="utf-8")
        else:
            html_text = get_with_retry(
                TGA_URL,
                attempts=TGA_HTML_ATTEMPTS,
                timeout=TGA_HTML_TIMEOUT,
            ).text
        frames = [flatten_columns(x) for x in pd.read_html(StringIO(html_text))]
        frame = next(
            (
                x
                for x in frames
                if any("nitrosamine" in normalize(c) for c in x.columns)
                and any("ailimit" in normalize(c) for c in x.columns)
            ),
            None,
        )
        if frame is None:
            raise ValueError(
                f"TGA AI table not found; CSV error was: {csv_error}"
            )
        soup = BeautifulSoup(html_text, "lxml")
        text = soup.get_text(" ", strip=True)
        match = re.search(r"Last updated\s+(\d{1,2}\s+\w+\s+\d{4})", text, flags=re.I)
        page_updated = iso_date(match.group(1)) if match else ""

    assert frame is not None
    name_col = find_column(frame, "Nitrosamine")
    ai_col = find_column(frame, "AI limit")
    source_col = next((c for c in frame.columns if normalize(c).startswith("source")), "")
    cpca_col = next((c for c in frame.columns if "cpca" in normalize(c)), "")
    cas_col = next((c for c in frame.columns if "casno" in normalize(c)), "")
    published_col = next(
        (c for c in frame.columns if any(x in normalize(c) for x in ("datepublished", "firstpublished"))),
        "",
    )
    updated_col = next((c for c in frame.columns if "date" in normalize(c) and "updated" in normalize(c)), "")

    parsed = 0
    row_dates: list[str] = []
    for _, row in frame.iterrows():
        raw_name = clean(row.get(name_col))
        if not raw_name or raw_name.lower() == "nan" or "default class specific" in raw_name.lower():
            continue
        name, embedded_cas, aliases = split_tga_name(raw_name)
        separate_cas, separate_aliases = split_tga_identifiers(row.get(cas_col)) if cas_col else ("", [])
        cas = separate_cas or embedded_cas
        aliases.extend(separate_aliases)
        publication_date = latest_tga_date(
            row.get(published_col) if published_col else "",
            row.get(updated_col) if updated_col else "",
        )
        if publication_date:
            row_dates.append(publication_date)
        registry.add_regulator(
            agency="TGA", name=name, cas=cas, aliases=aliases,
            related=[row.get(source_col)] if source_col else [],
            ai_raw=row.get(ai_col), cpca=row.get(cpca_col) if cpca_col else "",
            basis="Established AI", publication_date=publication_date,
            source_url=TGA_URL, source_version=f"TGA established AI table ({mode})",
            source_table="TGA established AI table"
        )
        parsed += 1

    if parsed < 50:
        raise ValueError(f"Only {parsed} TGA rows parsed; refusing a likely incomplete refresh")
    updated = page_updated or max(row_dates, default=dt.date.today().isoformat())
    return source_status(
        f"TGA established AI table via {mode}; {parsed} rows",
        updated,
        TGA_URL,
        mode=mode,
    )


def add_manual_regulatory_references(registry: Registry) -> None:
    """Add transparent guidance-derived references not listed as FDA rows.

    NDBA is retained as a 26.5 ng/day *reference* value because FDA's current
    online compound-specific tables do not list it as a dedicated row.  The
    status and basis make that distinction explicit in the UI.
    """
    record = registry.get(
        "N-nitroso-di-n-butylamine",
        "924-16-3",
        aliases=("NDBA", "N-nitrosodibutylamine", "N-亚硝基二正丁胺"),
    )
    if "FDA" not in record["regulators"]:
        registry.add_regulator(
            agency="FDA",
            name="N-nitroso-di-n-butylamine",
            cas="924-16-3",
            aliases=("NDBA", "N-nitrosodibutylamine", "N-亚硝基二正丁胺"),
            ai_raw=26.5,
            basis=(
                "FDA当前在线化合物特异性表未单列；此处为现行指导原则下的保守/默认参考值，"
                "申报时应核对具体适用性。"
            ),
            publication_date="2024-09-04",
            source_url=FDA_URL,
            source_version="Guidance-derived reference",
            status="reference",
            source_table="FDA guidance-derived reference",
        )


def load_aliases() -> dict[str, list[str]]:
    if not ALIASES_FILE.exists():
        return {}
    return json.loads(ALIASES_FILE.read_text(encoding="utf-8"))


def load_previous() -> dict[str, Any]:
    if DATA_FILE.exists():
        return json.loads(DATA_FILE.read_text(encoding="utf-8"))
    return {"records": [], "source_status": {}}


def seed_previous_agency(registry: Registry, previous: dict[str, Any], agency: str) -> None:
    for old in previous.get("records", []):
        item = old.get("regulators", {}).get(agency)
        special_items = [
            copy.deepcopy(value)
            for value in old.get("special_limits", [])
            if clean(value.get("agency")) == agency
        ]
        if not item and not special_items:
            continue
        record = registry.get(old["name"], old.get("cas", ""), old.get("aliases", []))
        record["related_substances"] = list(dict.fromkeys(record["related_substances"] + old.get("related_substances", [])))
        record["smiles"] = record["smiles"] or old.get("smiles", "")
        record["iupac"] = record["iupac"] or old.get("iupac", "")
        if item:
            record["regulators"][agency] = copy.deepcopy(item)
        record["special_limits"].extend(special_items)


def record_index(payload: dict[str, Any]) -> dict[tuple[str, str], dict[str, Any]]:
    result = {}
    for record in payload.get("records", []):
        key = (record.get("cas") or normalize(record.get("name")), record.get("name", ""))
        result[key] = record
    return result


def diff_payload(old: dict[str, Any], new: dict[str, Any]) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []
    old_by_cas = {r.get("cas"): r for r in old.get("records", []) if r.get("cas")}
    old_by_name = {normalize(r.get("name")): r for r in old.get("records", [])}
    matched_old_ids: set[int] = set()

    def special_identity(item: dict[str, Any]) -> tuple[str, str, str]:
        return (
            normalize(item.get("agency")),
            normalize(item.get("limit_type")),
            normalize(item.get("applicable_product")),
        )

    def special_summary(item: dict[str, Any] | None) -> str | None:
        if not item:
            return None
        bits = [clean(item.get("ai_display"))]
        if clean(item.get("official_control_display")):
            bits.append(f"control {clean(item.get('official_control_display'))}")
        if clean(item.get("estimated_duration")):
            bits.append(f"to {clean(item.get('estimated_duration'))}")
        return "; ".join(filter(None, bits))

    for new_record in new.get("records", []):
        old_record = old_by_cas.get(new_record.get("cas")) if new_record.get("cas") else None
        old_record = old_record or old_by_name.get(normalize(new_record.get("name")))
        if old_record is not None:
            matched_old_ids.add(id(old_record))

        if old_record is None:
            for agency, value in new_record.get("regulators", {}).items():
                changes.append({
                    "type": "added", "compound": new_record["name"], "cas": new_record.get("cas", ""),
                    "agency": agency, "old": None, "new": value.get("ai_display")
                })
            for value in new_record.get("special_limits", []):
                changes.append({
                    "type": "added", "compound": new_record["name"], "cas": new_record.get("cas", ""),
                    "agency": value.get("agency", ""), "scope": value.get("limit_type", "special"),
                    "old": None, "new": special_summary(value),
                })
            continue

        agencies = set(old_record.get("regulators", {})) | set(new_record.get("regulators", {}))
        for agency in agencies:
            before = old_record.get("regulators", {}).get(agency)
            after = new_record.get("regulators", {}).get(agency)
            if before is None and after is not None:
                changes.append({"type": "added", "compound": new_record["name"], "cas": new_record.get("cas", ""),
                                "agency": agency, "old": None, "new": after.get("ai_display")})
            elif before is not None and after is None:
                changes.append({"type": "removed", "compound": new_record["name"], "cas": new_record.get("cas", ""),
                                "agency": agency, "old": before.get("ai_display"), "new": None})
            elif before and after:
                fields = ("ai_display", "cpca_category", "basis", "source_table")
                if any(clean(before.get(field)) != clean(after.get(field)) for field in fields):
                    changes.append({"type": "changed", "compound": new_record["name"], "cas": new_record.get("cas", ""),
                                    "agency": agency, "old": before.get("ai_display"), "new": after.get("ai_display")})

        old_special = {special_identity(item): item for item in old_record.get("special_limits", [])}
        new_special = {special_identity(item): item for item in new_record.get("special_limits", [])}
        for identity in set(old_special) | set(new_special):
            before = old_special.get(identity)
            after = new_special.get(identity)
            agency = clean((after or before or {}).get("agency"))
            scope = clean((after or before or {}).get("limit_type")) or "special"
            if before is None:
                change_type = "added"
            elif after is None:
                change_type = "removed"
            else:
                fields = ("ai_display", "official_control_display", "estimated_duration", "basis")
                if not any(clean(before.get(field)) != clean(after.get(field)) for field in fields):
                    continue
                change_type = "changed"
            changes.append({
                "type": change_type,
                "compound": new_record["name"],
                "cas": new_record.get("cas", ""),
                "agency": agency,
                "scope": scope,
                "old": special_summary(before),
                "new": special_summary(after),
            })

    # Detect records that disappeared completely from all parsed sources.
    for old_record in old.get("records", []):
        if id(old_record) in matched_old_ids:
            continue
        for agency, value in old_record.get("regulators", {}).items():
            changes.append({
                "type": "removed", "compound": old_record["name"], "cas": old_record.get("cas", ""),
                "agency": agency, "old": value.get("ai_display"), "new": None,
            })
        for value in old_record.get("special_limits", []):
            changes.append({
                "type": "removed", "compound": old_record["name"], "cas": old_record.get("cas", ""),
                "agency": value.get("agency", ""), "scope": value.get("limit_type", "special"),
                "old": special_summary(value), "new": None,
            })
    return changes

def send_dingtalk(
    changes: list[dict[str, Any]],
    errors: dict[str, str] | None = None,
) -> None:
    errors = errors or {}
    beijing_now = dt.datetime.now(
        dt.timezone(dt.timedelta(hours=8))
    )
    is_monday = beijing_now.weekday() == 0

    # No notification is expected on an ordinary non-Monday run with no
    # regulatory changes and no source errors.
    if not changes and not errors and not is_monday:
        return

    url = os.getenv("DINGTALK_RELAY_URL", "").strip()
    token = os.getenv("DINGTALK_RELAY_TOKEN", "").strip()
    missing = [
        name
        for name, value in (
            ("DINGTALK_RELAY_URL", url),
            ("DINGTALK_RELAY_TOKEN", token),
        )
        if not value
    ]
    if missing:
        raise RuntimeError(
            "DingTalk notification configuration missing: " + ", ".join(missing)
        )

    run_time = beijing_now.strftime("%Y/%m/%d %H:%M")
    delivery_failures: list[str] = []

    def deliver(title: str, lines: list[str]) -> None:
        response = SESSION.post(
            url,
            timeout=30,
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            },
            json={"title": title, "message": "\n".join(lines)},
        )
        try:
            relay_result = response.json()
        except ValueError:
            relay_result = None
        if not response.ok:
            detail = relay_result if relay_result is not None else clean(response.text)[:1000]
            raise RuntimeError(
                f"Cloudflare relay HTTP {response.status_code}: {detail}"
            )
        if relay_result is None:
            raise RuntimeError("Cloudflare relay returned non-JSON response")
        if relay_result.get("ok") is not True:
            raise RuntimeError(f"Cloudflare relay rejected notification: {relay_result}")
        print(f"[OK] DingTalk delivered: {title}")

    def attempt_delivery(title: str, lines: list[str]) -> None:
        try:
            deliver(title, lines)
        except Exception as exc:
            message = f"{title}: {type(exc).__name__}: {exc}"
            delivery_failures.append(message)
            print(f"[WARN] DingTalk delivery failed: {message}", file=sys.stderr)

    # Regulatory changes are always sent as their own message.
    if changes:
        title = "亚硝胺监管限度更新"
        lines = [
            f"### 亚硝胺监管限度更新｜{beijing_now:%Y/%m/%d}",
            f"运行时间：{run_time}（北京时间）",
            "监测范围：FDA、EMA、Health Canada、TGA",
            "",
            f"本次识别到 **{len(changes)}** 项变化：",
        ]

        for change in changes[:30]:
            compound = change.get("compound") or "未命名化合物"
            cas_text = f" ({change['cas']})" if change.get("cas") else ""
            scope_text = f" ({change['scope']})" if change.get("scope") else ""

            lines.append(
                f"- **{compound}**{cas_text}｜"
                f"{change.get('agency', '未知机构')}{scope_text}｜"
                f"{change.get('type', '更新')}："
                f"{change.get('old') or '—'} → {change.get('new') or '—'}"
            )

        if len(changes) > 30:
            lines.append(
                f"- 另有 {len(changes) - 30} 项变化，"
                "请打开查询网页或 changes.json 查看。"
            )
        attempt_delivery(title, lines)

    # Source errors are always sent separately, so a failed regulator cannot
    # be hidden inside the Monday heartbeat or a regulatory-change message.
    if errors:
        title = "亚硝胺监管监测异常"
        lines = [
            f"### ⚠ 亚硝胺监管监测异常｜{beijing_now:%Y/%m/%d}",
            f"运行时间：{run_time}（北京时间）",
            "监测范围：FDA、EMA、Health Canada、TGA",
            "",
            "本次未能完整读取全部监管机构数据。",
        ]
        lines.extend(["", "### ⚠ 数据源读取异常"])
        for agency, error in errors.items():
            lines.append(f"- **{agency}**：{error}")
        lines.append(
            "- 系统已保留读取失败机构的上一版数据，"
            "建议结合官方原始文件进行人工核验。"
        )
        attempt_delivery(title, lines)

    # Monday heartbeat is independent: it is sent even when the same run also
    # produced a change notification and/or a source-error notification.
    if is_monday:
        title = "亚硝胺监管监测周报"
        status_text = (
            f"⚠ 本次运行存在 {len(errors)} 个数据源异常。"
            if errors
            else "✅ 本次监测运行正常。"
        )
        lines = [
            f"### 亚硝胺监管监测周报｜{beijing_now:%Y/%m/%d}",
            f"运行时间：{run_time}（北京时间）",
            "监测范围：FDA、EMA、Health Canada、TGA",
            "",
            status_text,
            f"本次识别到 {len(changes)} 项监管数据变化。",
            "如有变化或数据源异常，系统已另行发送独立消息。",
        ]
        attempt_delivery(title, lines)

    if delivery_failures:
        raise RuntimeError("; ".join(delivery_failures))


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--allow-partial", action="store_true", help="Keep previous data when a source fails")
    parser.add_argument("--no-notify", action="store_true")
    args = parser.parse_args()

    previous = load_previous()
    registry = Registry(load_aliases())
    statuses: dict[str, dict[str, str]] = {}
    errors: dict[str, str] = {}

    sources = [
        ("EMA", lambda: parse_ema(Path(os.environ["EMA_XLSX_PATH"]) if os.getenv("EMA_XLSX_PATH") else download(EMA_URL, ".xlsx"), registry)),
        ("Health Canada", lambda: parse_health_canada(Path(os.environ["HC_XLSX_PATH"]) if os.getenv("HC_XLSX_PATH") else download(HC_URL, ".xlsx"), registry)),
        ("FDA", lambda: parse_fda(registry)),
        ("TGA", lambda: parse_tga(registry)),
    ]

    for agency, operation in sources:
        try:
            statuses[agency] = operation()
            print(f"[OK] {agency}: {statuses[agency]['version']}")
        except Exception as exc:  # network/parsing failures must not erase good historical data
            errors[agency] = f"{type(exc).__name__}: {exc}"
            print(f"[ERROR] {agency}: {errors[agency]}", file=sys.stderr)
            if not args.allow_partial:
                raise
            seed_previous_agency(registry, previous, agency)
            old_status = previous.get("source_status", {}).get(agency, {})
            statuses[agency] = {
                **old_status,
                "mode": f"previous data retained; refresh failed: {errors[agency]}"
            }

    add_manual_regulatory_references(registry)
    registry.apply_aliases()
    generated_at = dt.datetime.now(dt.timezone.utc).astimezone().isoformat(timespec="seconds")
    payload = {
        "schema_version": 2,
        "generated_at": generated_at,
        "disclaimer": "本数据库用于研发和法规检索辅助。AI为每日可接受摄入量，ppm必须按最大日剂量换算；申报和放行前应打开监管机构原始文件核验。",
        "source_status": statuses,
        "refresh_errors": errors,
        "records": registry.serialise(),
    }
    changes = diff_payload(previous, payload)

    DATA_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    CHANGES_FILE.write_text(json.dumps({"generated_at": generated_at, "changes": changes}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(payload['records'])} records; {len(changes)} changes.")

    # Bootstrap data is only an installation baseline. When FDA/TGA are expanded
    # to their full official tables on the first GitHub Actions run, do not flood
    # DingTalk with hundreds of baseline additions. Subsequent changes notify normally.
    bootstrap_agencies = {
        agency
        for agency, item in previous.get("source_status", {}).items()
        if "bootstrap" in clean(item.get("mode")).lower()
    }
    notification_changes = [
        change for change in changes if change.get("agency") not in bootstrap_agencies
    ]

    if not args.no_notify:
        try:
            send_dingtalk(notification_changes, errors)
        except Exception as exc:
            print(f"[WARN] DingTalk notification failed: {exc}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
